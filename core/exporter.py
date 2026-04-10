"""
Governance-Suite — Exportación de resultados
Soporta CSV, Excel y JSON.
"""
import csv
import json
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from core.logger import get_logger
from config import OUTPUT_DIR

logger = get_logger("exporter")


def _resolve_path(filename: str, folder: Optional[Path] = None) -> Path:
    base = folder or OUTPUT_DIR
    base.mkdir(parents=True, exist_ok=True)
    return base / filename


def export_csv(data: List[Dict], filename: str, folder: Optional[Path] = None) -> str:
    """Exporta lista de dicts a CSV."""
    if not data:
        logger.warning("export_csv: datos vacíos")
        return ""
    path = _resolve_path(filename, folder)
    fields = list(data[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)
    logger.info(f"CSV exportado: {path}")
    return str(path)


def export_json(data, filename: str, folder: Optional[Path] = None) -> str:
    """Exporta datos a JSON."""
    path = _resolve_path(filename, folder)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"JSON exportado: {path}")
    return str(path)


def export_excel(data: List[Dict], filename: str, sheet_name: str = "Datos", folder: Optional[Path] = None) -> str:
    """Exporta lista de dicts a Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl no instalado. Instala con: pip install openpyxl")
        return ""

    if not data:
        logger.warning("export_excel: datos vacíos")
        return ""

    path = _resolve_path(filename, folder)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Cabeceras con estilo
    headers = list(data[0].keys())
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, row in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    # Autoajuste de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)
    logger.info(f"Excel exportado: {path}")
    return str(path)


def auto_export(data, base_name: str = None, fmt: str = "csv") -> str:
    """Exporta con nombre automático basado en timestamp."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{base_name or 'export'}_{ts}.{fmt}"
    if fmt == "csv":
        return export_csv(data, name)
    elif fmt == "json":
        return export_json(data, name)
    elif fmt in ("excel", "xlsx"):
        return export_excel(data, name.replace(".excel", ".xlsx"))
    else:
        logger.error(f"Formato desconocido: {fmt}")
        return ""
